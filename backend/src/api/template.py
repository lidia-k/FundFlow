"""Template API endpoint for downloading Excel template."""

from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse

router = APIRouter()


@router.get("/template/salt-rules")
async def download_salt_rules_template() -> FileResponse:
    """
    Download SALT rules matrix template file.

    Returns Excel template with Withholding and Composite sheets for SALT tax rules.
    """
    # Look for SALT rules template file in data directory
    salt_template_path = Path("data/templates/salt_matrix_template.xlsx")

    # Check if template exists
    if not salt_template_path.exists():
        raise HTTPException(
            status_code=404, detail="SALT rules template file not found"
        )

    # Return file
    return FileResponse(
        path=str(salt_template_path),
        filename="salt_rules_matrix_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=salt_rules_matrix_template.xlsx"
        },
    )


@router.get("/template")
async def download_template() -> FileResponse:
    """
    Download Excel template file for investor data upload.

    Returns Excel template with proper headers and formatting.
    """
    # Look for template file in data directory
    template_path = Path("data/templates/investor_data_template.xlsx")

    # Check if template exists
    if not template_path.exists():
        # If template doesn't exist, create a basic one
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        # Create template directory if it doesn't exist
        template_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Investor Data"

        # Define headers (v1.3 format)
        headers = [
            "Investor Name",
            "Investor Entity Type",
            "Investor Tax State",
            "Commitment Percentage",
            "Distribution TX",
            "Distribution NM",
            "Distribution CO",
            "NM Withholding Exemption",
            "NM Composite Exemption",
            "CO Composite Exemption",
            "CO Withholding Exemption",
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add example data row
        example_data = [
            "Example LP 1",
            "LLC_Taxed as Partnership",
            "NY",
            "10%",
            "100,000",
            "5,000",
            "50,000",
            "",
            "",
            "X",
            "",
        ]

        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            "FundFlow Investor Data Template Instructions (v1.3 Format)",
            "",
            "SHEET 1: Investor Data - Column Descriptions:",
            "• Investor Name: Full legal name of the investor entity",
            "• Investor Entity Type: LLC_Taxed as Partnership, Corporation, Individual, Trust, etc.",
            "• Investor Tax State: Two-letter state code (e.g., NY, CA, TX)",
            "• Commitment Percentage: Percentage with % symbol (e.g., 10%, 15%)",
            "• Distribution TX: Distribution amount for Texas",
            "• Distribution NM: Distribution amount for New Mexico",
            "• Distribution CO: Distribution amount for Colorado",
            "• NM Withholding Exemption: Enter 'X' if exempt, leave blank if not",
            "• NM Composite Exemption: Enter 'X' if exempt, leave blank if not",
            "• CO Composite Exemption: Enter 'X' if exempt, leave blank if not",
            "• CO Withholding Exemption: Enter 'X' if exempt, leave blank if not",
            "",
            "SHEET 2: Fund Source Data - Column Descriptions:",
            "• Company: Portfolio company name (e.g., Company A, Company B)",
            "• State: Two-letter state code where income is sourced (e.g., TX, NM, CO)",
            "• Share (%): Fund's ownership percentage in the company (0-100, no % symbol)",
            "• Distribution Amount: Total distribution amount for this company/state combination",
            "",
            "File Requirements:",
            "• Maximum file size: 10MB",
            "• Supported formats: .xlsx, .xls",
            "• Data should start from row 2 (headers in row 1) on both sheets",
            "• Fund Source Data sheet is optional for backward compatibility",
            "",
            "Validation Rules:",
            "• INVESTOR DATA: All fields are required except exemption fields",
            "• FUND SOURCE DATA: All columns are required when sheet is present",
            "• Distribution amounts should be in numeric format (can include commas)",
            "• State codes must be valid US states",
            "• Entity types must match allowed values",
            "• Commitment percentage must include % symbol",
            "• Fund share percentage must be 0-100 (without % symbol)",
            "• No duplicate Company/State combinations allowed in Fund Source Data",
            "• Exemption fields: use 'X' for exempt, leave blank for not exempt",
            "",
            "For support, contact: support@fundflow.com",
        ]

        for row, instruction in enumerate(instructions, 1):
            instructions_ws.cell(row=row, column=1, value=instruction)

        # Style instructions header
        instructions_ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Auto-adjust instructions column width
        instructions_ws.column_dimensions["A"].width = 80

        # Add Fund Source Data sheet
        fund_source_ws = wb.create_sheet("Fund Source Data")
        fund_source_ws.title = "Fund Source Data"

        # Fund Source Data headers
        fund_source_headers = ["Company", "State", "Share (%)", "Distribution Amount"]

        # Add headers with same styling
        for col, header in enumerate(fund_source_headers, 1):
            cell = fund_source_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add example fund source data
        fund_source_example = [
            ["Company A", "TX", "57.19", "10,000,000"],
            ["Company A", "NM", "28.18", "10,000,000"],
            ["Company B", "CO", "100.00", "5,000,000"],
        ]

        for row_idx, row_data in enumerate(fund_source_example, 2):
            for col, value in enumerate(row_data, 1):
                fund_source_ws.cell(row=row_idx, column=col, value=value)

        # Auto-adjust Fund Source Data column widths
        for column in fund_source_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            fund_source_ws.column_dimensions[column_letter].width = adjusted_width

        # Save template
        wb.save(str(template_path))

    # Verify template file exists
    if not template_path.exists():
        raise HTTPException(
            status_code=500, detail="Template file could not be created or found"
        )

    # Return file
    return FileResponse(
        path=str(template_path),
        filename="fundflow_investor_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=fundflow_investor_template.xlsx"
        },
    )
