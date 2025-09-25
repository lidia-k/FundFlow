"""Excel file validation and parsing service."""

import math
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook

from ..models.enums import InvestorEntityType, USJurisdiction
from ..models.validation_error import ErrorSeverity


class ExcelValidationError:
    """Represents a validation error found in Excel processing."""

    def __init__(
        self,
        row_number: int,
        column_name: str,
        error_code: str,
        error_message: str,
        severity: ErrorSeverity,
        field_value: str | None = None,
    ):
        self.row_number = row_number
        self.column_name = column_name
        self.error_code = error_code
        self.error_message = error_message
        self.severity = severity
        self.field_value = field_value


class ExcelParsingResult:
    """Result of Excel parsing operation."""

    def __init__(
        self,
        data: list[dict[str, Any]],
        errors: list[ExcelValidationError],
        fund_info: dict[str, str],
        total_rows: int,
        valid_rows: int,
        fund_source_data: list[dict[str, Any]] | None = None,
        fund_source_errors: list[ExcelValidationError] | None = None,
    ):
        self.data = data
        self.errors = errors
        self.fund_info = fund_info
        self.total_rows = total_rows
        self.valid_rows = valid_rows
        self.fund_source_data = fund_source_data or []
        self.fund_source_errors = fund_source_errors or []


class ExcelService:
    """Service for Excel file validation and parsing (v1.3 format)."""

    # Fixed base column headers for v1.3 format
    REQUIRED_BASE_HEADERS = [
        "Investor Name",
        "Investor Entity Type",
        "Investor Tax State",
        "Commitment Percentage",
    ]

    # Fund Source Data sheet required headers
    REQUIRED_FUND_SOURCE_HEADERS = [
        "Company",
        "State",
        "Share (%)",
        "Distribution",
    ]

    # Valid entity types
    VALID_ENTITY_TYPES = {entity_type.value for entity_type in InvestorEntityType}

    # Valid US state codes
    VALID_STATE_CODES = {state.value for state in USJurisdiction}

    def __init__(self):
        self.errors: list[ExcelValidationError] = []
        self.detected_columns: dict[str, dict[str, str]] = {
            "distribution": {},
            "withholding_exemption": {},
            "composite_exemption": {},
        }
        self._seen_investors = set()

    def _is_empty_value(self, value: Any) -> bool:
        """Return True when a cell value is considered empty."""
        if value is None:
            return True
        if isinstance(value, float) and math.isnan(value):
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    def _load_sheet_rows(
        self, workbook: Workbook, sheet_index: int
    ) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
        """Load a worksheet into normalized headers and row dictionaries."""
        worksheets = workbook.worksheets
        if sheet_index >= len(worksheets):
            raise IndexError("Worksheet index %s is invalid" % sheet_index)

        sheet = worksheets[sheet_index]
        header_row = next(
            sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
        )
        if not header_row:
            return [], []

        normalized_headers = [self.normalize_header(cell) for cell in header_row]
        rows: list[tuple[int, dict[str, Any]]] = []

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            row_dict: dict[str, Any] = {}
            for idx, header in enumerate(normalized_headers):
                value = row[idx] if idx < len(row) else None
                row_dict[header] = value
            rows.append((row_number, row_dict))

        return normalized_headers, rows

    def extract_fund_info_from_filename(
        self, filename: str
    ) -> dict[str, str] | None:
        """Extract fund code, quarter, and year from filename."""
        # v1.3 format pattern - uploaded as XLSX but following CSV naming convention
        pattern = r"^\(Input Data\) (.+)_Q([1-4]) (\d{4}) distribution data_v[\d\.]+\.(xlsx|xls)$"
        match = re.match(pattern, filename)

        if match:
            fund_code, quarter, year, extension = match.groups()
            return {
                "fund_code": fund_code.strip(),
                "period_quarter": f"Q{quarter}",
                "period_year": year,
                "file_extension": extension,
            }

        return None

    def validate_file_size(self, file_path: Path) -> bool:
        """Validate file size (max 10MB)."""
        file_size = file_path.stat().st_size
        max_size = 10 * 1024 * 1024  # 10MB

        if file_size > max_size:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="file",
                    error_code="FILE_SIZE_EXCEEDED",
                    error_message=f"File size {file_size} bytes exceeds 10MB limit",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return False
        return True

    def normalize_header(self, header: str) -> str:
        """Normalize header by trimming and collapsing whitespace."""
        return re.sub(r"\s+", " ", str(header).strip())

    def detect_dynamic_columns(self, headers: list[str]) -> bool:
        """Detect distribution, withholding exemption, and composite exemption columns by pattern."""
        normalized_headers = [self.normalize_header(col) for col in headers]

        # Reset detected columns
        self.detected_columns = {
            "distribution": {},
            "withholding_exemption": {},
            "composite_exemption": {},
        }

        for header in normalized_headers:
            # Distribution pattern: "Distribution" + space + state (e.g., "Distribution TX")
            if header.startswith("Distribution ") and len(header) == len(
                "Distribution XX"
            ):
                state = header.split(" ")[1].upper()
                if state in self.VALID_STATE_CODES:
                    self.detected_columns["distribution"][state] = header
                continue

            # Withholding Exemption pattern: state + space + "Withholding Exemption"
            if (
                header.endswith(" Withholding Exemption")
                and len(header.split(" ")) == 3
            ):
                state = header.split(" ")[0].upper()
                if state in self.VALID_STATE_CODES:
                    self.detected_columns["withholding_exemption"][state] = header
                continue

            # Composite Exemption pattern: state + "Composite Exemption" (with or without space)
            # Handles both "CO Composite Exemption" and "CO CompositeExemption"
            if (
                header.endswith(" Composite Exemption") and len(header.split(" ")) == 3
            ) or (
                header.endswith("CompositeExemption") and len(header.split(" ")) == 2
            ):
                state = header.split(" ")[0].upper()
                if state in self.VALID_STATE_CODES:
                    self.detected_columns["composite_exemption"][state] = header
                continue

        # Check if we found at least one distribution column
        if not self.detected_columns["distribution"]:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="distribution_columns",
                    error_code="NO_DISTRIBUTION_COLUMNS",
                    error_message="No distribution columns found. Expected format: 'Distribution XX' where XX is state code",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return False

        return True

    def validate_headers(self, headers: list[str]) -> bool:
        """Validate that all required headers are present."""
        normalized_headers = [self.normalize_header(col) for col in headers]

        # Detect dynamic columns for distribution and exemptions
        if not self.detect_dynamic_columns(headers):
            return False

        # Check required base headers (fixed headers)
        missing_headers = []
        normalized_required = [
            self.normalize_header(col) for col in self.REQUIRED_BASE_HEADERS
        ]

        for required_header in normalized_required:
            if required_header not in normalized_headers:
                missing_headers.append(required_header)

        if missing_headers:
            for header in missing_headers:
                self.errors.append(
                    ExcelValidationError(
                        row_number=0,
                        column_name=header,
                        error_code="MISSING_HEADER",
                        error_message=f"Required column header '{header}' is missing",
                        severity=ErrorSeverity.ERROR,
                    )
                )
            return False
        return True

    def parse_numeric_value(
        self, value: Any, column_name: str, row_num: int
    ) -> Decimal:
        """Parse numeric value with thousands separators and parentheses."""
        if self._is_empty_value(value):
            return Decimal("0.00")

        str_value = str(value).strip()

        # Handle parentheses (negative values - which are invalid)
        if str_value.startswith("(") and str_value.endswith(")"):
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name=column_name,
                    error_code="NEGATIVE_AMOUNT",
                    error_message=f"Negative amount {str_value} is not allowed",
                    severity=ErrorSeverity.ERROR,
                    field_value=str_value,
                )
            )
            return Decimal("0.00")

        # Remove thousands separators and spaces
        cleaned_value = re.sub(r"[,\s]", "", str_value)

        try:
            return Decimal(cleaned_value)
        except (InvalidOperation, ValueError):
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name=column_name,
                    error_code="INVALID_NUMBER_FORMAT",
                    error_message=f"Invalid number format: {str_value}",
                    severity=ErrorSeverity.ERROR,
                    field_value=str_value,
                )
            )
            return Decimal("0.00")

    def parse_exemption_value(self, value: Any) -> bool:
        """Parse exemption field value to boolean."""
        if self._is_empty_value(value):
            return False

        str_value = str(value).strip().lower()
        return str_value in ["exemption", "x", "true", "yes", "1"]

    def parse_percentage_value(self, value: Any) -> Decimal | None:
        """Parse percentage value for fund source data (simple version)."""
        if self._is_empty_value(value):
            return None

        str_value = str(value).strip()
        # Remove % symbol if present
        str_value = str_value.replace("%", "")

        try:
            return Decimal(str_value)
        except (InvalidOperation, ValueError):
            return None

    def _parse_and_validate_commitment_percentage(
        self,
        value: Any,
        row_num: int
    ) -> Optional[Decimal]:
        """Parse commitment percentage ensuring required format and bounds."""
        if self._is_empty_value(value) or str(value).strip() == "":
            self.errors.append(ExcelValidationError(
                row_number=row_num,
                column_name='Commitment Percentage',
                error_code="EMPTY_FIELD",
                error_message="Commitment Percentage is required",
                severity=ErrorSeverity.ERROR
            ))
            return None

        str_value = str(value).strip().replace('%', '')

        try:
            decimal_value = Decimal(str_value)
        except (InvalidOperation, ValueError):
            self.errors.append(ExcelValidationError(
                row_number=row_num,
                column_name='Commitment Percentage',
                error_code="INVALID_PERCENTAGE_FORMAT",
                error_message=f"Invalid percentage format: {str_value}",
                severity=ErrorSeverity.ERROR,
                field_value=str_value
            ))
            return None

        normalized_value = decimal_value.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)

        if normalized_value < Decimal('0') or normalized_value > Decimal('100'):
            self.errors.append(ExcelValidationError(
                row_number=row_num,
                column_name='Commitment Percentage',
                error_code="PERCENTAGE_OUT_OF_RANGE",
                error_message="Commitment Percentage must be between 0 and 100",
                severity=ErrorSeverity.ERROR,
                field_value=str(normalized_value)
            ))
            return None

        return normalized_value

    def validate_base_fields(self, row_data: dict[str, Any], row_num: int) -> bool:
        """Validate common fields present in both formats."""
        is_valid = True

        # Validate investor name
        investor_name = str(row_data.get("Investor Name", "")).strip()
        if not investor_name:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Investor Name",
                    error_code="EMPTY_FIELD",
                    error_message="Investor Name cannot be empty",
                    severity=ErrorSeverity.ERROR,
                )
            )
            is_valid = False

        # Validate entity type
        entity_type = str(row_data.get("Investor Entity Type", "")).strip()
        if entity_type not in self.VALID_ENTITY_TYPES:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Investor Entity Type",
                    error_code="INVALID_ENTITY_TYPE",
                    error_message=f"Invalid entity type: {entity_type}",
                    severity=ErrorSeverity.ERROR,
                    field_value=entity_type,
                )
            )
            is_valid = False

        # Validate tax state
        tax_state = str(row_data.get("Investor Tax State", "")).strip().upper()
        if tax_state not in self.VALID_STATE_CODES:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Investor Tax State",
                    error_code="INVALID_STATE_CODE",
                    error_message=f"Invalid state code: {tax_state}",
                    severity=ErrorSeverity.ERROR,
                    field_value=tax_state,
                )
            )
            is_valid = False

        parsed_commitment = self._parse_and_validate_commitment_percentage(
            row_data.get('Commitment Percentage'),
            row_num
        )
        if parsed_commitment is None:
            is_valid = False
        else:
            row_data['_parsed_commitment_percentage'] = parsed_commitment

        return is_valid

    def validate_row_data(self, row_data: dict[str, Any], row_num: int) -> bool:
        """Validate row data for v1.3 format."""
        is_valid = self.validate_base_fields(row_data, row_num)

        if is_valid:
            investor_key = (
                str(row_data.get('Investor Name', '')).strip().lower(),
                str(row_data.get('Investor Entity Type', '')).strip(),
                str(row_data.get('Investor Tax State', '')).strip().upper()
            )
            if investor_key in self._seen_investors:
                self.errors.append(ExcelValidationError(
                    row_number=row_num,
                    column_name='Investor Name',
                    error_code="DUPLICATE_INVESTOR",
                    error_message="Duplicate investor rows detected in upload",
                    severity=ErrorSeverity.ERROR
                ))
                is_valid = False
            else:
                self._seen_investors.add(investor_key)

        # Check for at least one distribution amount > 0
        has_distribution = False
        for state, col_name in self.detected_columns["distribution"].items():
            amount = self.parse_numeric_value(row_data.get(col_name), col_name, row_num)
            if amount > 0:
                has_distribution = True
                break

        if not has_distribution:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Distribution Amounts",
                    error_code="ZERO_DISTRIBUTIONS",
                    error_message="At least one distribution amount must be greater than 0",
                    severity=ErrorSeverity.ERROR,
                )
            )
            is_valid = False

        return is_valid

    def parse_row(self, row_data: dict[str, Any], row_num: int) -> dict[str, Any]:
        """Parse row data for v1.3 format."""
        parsed_row = {
            "investor_name": str(row_data["Investor Name"]).strip(),
            "investor_entity_type": str(row_data["Investor Entity Type"]).strip(),
            "investor_tax_state": str(row_data["Investor Tax State"]).strip().upper(),
            "row_number": row_num,
        }

        # Parse commitment percentage (use validated value from validation step)
        parsed_row['commitment_percentage'] = row_data.get('_parsed_commitment_percentage')

        # Parse distribution amounts by state
        parsed_row["distributions"] = {}
        for state, col_name in self.detected_columns["distribution"].items():
            amount = self.parse_numeric_value(row_data.get(col_name), col_name, row_num)
            parsed_row["distributions"][state] = amount

        # Parse withholding exemptions by state
        parsed_row["withholding_exemptions"] = {}
        for state, col_name in self.detected_columns["withholding_exemption"].items():
            exemption = self.parse_exemption_value(row_data.get(col_name))
            parsed_row["withholding_exemptions"][state] = exemption

        # Parse composite exemptions by state
        parsed_row["composite_exemptions"] = {}
        for state, col_name in self.detected_columns["composite_exemption"].items():
            exemption = self.parse_exemption_value(row_data.get(col_name))
            parsed_row["composite_exemptions"][state] = exemption

        return parsed_row

    def validate_fund_source_headers(self, headers: list[str]) -> bool:
        """Validate that all required Fund Source Data headers are present."""
        normalized_headers = [self.normalize_header(col) for col in headers]

        missing_headers = []
        normalized_required = [
            self.normalize_header(col) for col in self.REQUIRED_FUND_SOURCE_HEADERS
        ]

        for required_header in normalized_required:
            if required_header not in normalized_headers:
                missing_headers.append(required_header)

        if missing_headers:
            for header in missing_headers:
                self.errors.append(
                    ExcelValidationError(
                        row_number=0,
                        column_name=header,
                        error_code="MISSING_FUND_SOURCE_HEADER",
                        error_message=f"Required Fund Source Data column header '{header}' is missing",
                        severity=ErrorSeverity.ERROR,
                    )
                )
            return False
        return True

    def validate_fund_source_row(self, row_data: dict[str, Any], row_num: int) -> bool:
        """Validate fund source data row."""
        is_valid = True

        # Validate company name
        company_name = str(row_data.get("Company", "")).strip()
        if not company_name:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Company",
                    error_code="EMPTY_COMPANY_NAME",
                    error_message="Company name cannot be empty",
                    severity=ErrorSeverity.ERROR,
                )
            )
            is_valid = False

        # Validate state code
        state_code = str(row_data.get("State", "")).strip().upper()
        if state_code not in self.VALID_STATE_CODES:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="State",
                    error_code="INVALID_STATE_CODE",
                    error_message=f"Invalid state code: {state_code}",
                    severity=ErrorSeverity.ERROR,
                    field_value=state_code,
                )
            )
            is_valid = False

        # Validate share percentage (0-100)
        share_pct = self.parse_percentage_value(row_data.get("Share (%)"))
        if share_pct is None:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Share (%)",
                    error_code="INVALID_SHARE_PERCENTAGE",
                    error_message="Invalid share percentage format",
                    severity=ErrorSeverity.ERROR,
                    field_value=str(row_data.get("Share (%)")),
                )
            )
            is_valid = False
        elif share_pct < 0 or share_pct > 100:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Share (%)",
                    error_code="SHARE_PERCENTAGE_OUT_OF_RANGE",
                    error_message=f"Share percentage must be between 0 and 100, got {share_pct}",
                    severity=ErrorSeverity.ERROR,
                    field_value=str(share_pct),
                )
            )
            is_valid = False

        # Validate distribution amount (must be positive)
        dist_amount = self.parse_numeric_value(
            row_data.get("Distribution"), "Distribution", row_num
        )
        if dist_amount <= 0:
            self.errors.append(
                ExcelValidationError(
                    row_number=row_num,
                    column_name="Distribution",
                    error_code="INVALID_DISTRIBUTION_AMOUNT",
                    error_message="Distribution amount must be positive",
                    severity=ErrorSeverity.ERROR,
                    field_value=str(row_data.get("Distribution")),
                )
            )
            is_valid = False

        return is_valid

    def parse_fund_source_row(
        self, row_data: dict[str, Any], row_num: int
    ) -> dict[str, Any]:
        """Parse fund source data row."""
        parsed_row = {
            "company_name": str(row_data["Company"]).strip(),
            "state_jurisdiction": str(row_data["State"]).strip().upper(),
            "fund_share_percentage": self.parse_percentage_value(row_data["Share (%)"]),
            "total_distribution_amount": self.parse_numeric_value(
                row_data["Distribution"], "Distribution", row_num
            ),
            "row_number": row_num,
        }
        return parsed_row

    def parse_fund_source_data(
        self, workbook: Workbook
    ) -> tuple[list[dict[str, Any]], list[ExcelValidationError]]:
        """Parse Fund Source Data from second sheet."""
        fund_source_errors: list[ExcelValidationError] = []

        try:
            headers, rows = self._load_sheet_rows(workbook, 1)

            # Remove completely empty rows by checking if any cell has data
            filtered_rows = [
                (row_num, row)
                for row_num, row in rows
                if any(not self._is_empty_value(value) for value in row.values())
            ]

            if not filtered_rows:
                return [], []

            # Validate headers
            if not self.validate_fund_source_headers(headers):
                # Move errors from self.errors to fund_source_errors
                header_errors = self.errors[:]
                self.errors = []  # Clear service errors since we're returning them separately
                return [], header_errors

            valid_fund_source_data: list[dict[str, Any]] = []
            company_state_combinations = set()

            for row_num, row_data in filtered_rows:
                if self.validate_fund_source_row(row_data, row_num):
                    parsed_row = self.parse_fund_source_row(row_data, row_num)
                    combo_key = (
                        parsed_row["company_name"],
                        parsed_row["state_jurisdiction"],
                    )

                    if combo_key in company_state_combinations:
                        fund_source_errors.append(
                            ExcelValidationError(
                                row_number=row_num,
                                column_name="Company/State",
                                error_code="DUPLICATE_COMPANY_STATE",
                                error_message=(
                                    f"Duplicate Company/State combination: {combo_key[0]}/{combo_key[1]}"
                                ),
                                severity=ErrorSeverity.ERROR,
                            )
                        )
                    else:
                        company_state_combinations.add(combo_key)
                        valid_fund_source_data.append(parsed_row)

            return valid_fund_source_data, fund_source_errors
        except IndexError:
            return [], []  # Missing sheet is acceptable for backward compatibility
        except Exception as e:  # pragma: no cover - defensive guard
            fund_source_errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="fund_source_sheet",
                    error_code="FUND_SOURCE_PARSING_ERROR",
                    error_message=f"Failed to parse Fund Source Data sheet: {str(e)}",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return [], fund_source_errors

    def parse_excel_file(
        self, file_path: Path, original_filename: str
    ) -> ExcelParsingResult:
        """Parse Excel file and validate data (v1.3 format)."""
        self.errors = []  # Reset errors
        self._seen_investors = set()

        # Validate file size
        if not self.validate_file_size(file_path):
            return ExcelParsingResult([], self.errors, {}, 0, 0)

        # Extract fund info from filename
        fund_info = self.extract_fund_info_from_filename(original_filename)
        if not fund_info:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="filename",
                    error_code="INVALID_FILENAME",
                    error_message=f"Filename '{original_filename}' does not match required pattern",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return ExcelParsingResult([], self.errors, {}, 0, 0)

        try:
            workbook = load_workbook(filename=file_path, data_only=True)
        except InvalidFileException as exc:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="file",
                    error_code="FAILED_PARSING",
                    error_message=f"Failed to open Excel file: {str(exc)}",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return ExcelParsingResult([], self.errors, fund_info, 0, 0)
        except Exception as exc:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="file",
                    error_code="FAILED_PARSING",
                    error_message=f"Failed to open Excel file: {str(exc)}",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return ExcelParsingResult([], self.errors, fund_info, 0, 0)

        try:
            headers, rows = self._load_sheet_rows(workbook, 0)

            # Remove rows where Investor Name is empty or whitespace
            filtered_rows = [
                (row_num, row)
                for row_num, row in rows
                if not self._is_empty_value(row.get("Investor Name"))
            ]

            total_rows = len(filtered_rows)

            # Check row limit
            if total_rows > 50000:
                self.errors.append(
                    ExcelValidationError(
                        row_number=0,
                        column_name="file",
                        error_code="ROW_LIMIT_EXCEEDED",
                        error_message=f"File has {total_rows} rows, exceeding 50,000 row limit",
                        severity=ErrorSeverity.ERROR,
                    )
                )
                return ExcelParsingResult([], self.errors, fund_info, total_rows, 0)

            # Validate headers
            if not self.validate_headers(headers):
                return ExcelParsingResult([], self.errors, fund_info, total_rows, 0)

            valid_data: list[dict[str, Any]] = []
            valid_row_count = 0

            for row_num, row_data in filtered_rows:
                if self.validate_row_data(row_data, row_num):
                    parsed_row = self.parse_row(row_data, row_num)
                    valid_data.append(parsed_row)
                    valid_row_count += 1

            # Parse Fund Source Data from second sheet (optional)
            fund_source_data, fund_source_errors = self.parse_fund_source_data(workbook)

            # Combine all errors
            all_errors = self.errors + fund_source_errors

            return ExcelParsingResult(
                data=valid_data,
                errors=all_errors,
                fund_info=fund_info,
                total_rows=total_rows,
                valid_rows=valid_row_count,
                fund_source_data=fund_source_data,
                fund_source_errors=fund_source_errors,
            )
        except IndexError:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="worksheet",
                    error_code="MISSING_WORKSHEET",
                    error_message="Primary worksheet is missing",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return ExcelParsingResult([], self.errors, fund_info, 0, 0)
        except Exception as e:
            self.errors.append(
                ExcelValidationError(
                    row_number=0,
                    column_name="file",
                    error_code="FAILED_PARSING",
                    error_message=f"Failed to parse Excel file: {str(e)}",
                    severity=ErrorSeverity.ERROR,
                )
            )
            return ExcelParsingResult([], self.errors, fund_info, 0, 0)
