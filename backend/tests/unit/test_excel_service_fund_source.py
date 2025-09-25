"""Unit tests for ExcelService fund source data parsing functionality."""

from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest

from src.services.excel_service import ExcelService


class TestExcelServiceFundSource:
    """Test Excel service fund source data parsing methods."""

    def setup_method(self):
        """Set up test fixtures."""
        self.excel_service = ExcelService()
        self.test_data_dir = Path(__file__).parent.parent / "test_data"

    def test_validate_fund_source_headers_valid(self):
        """Test fund source header validation with valid headers."""
        df = pd.DataFrame(
            {
                "Company": ["TechCorp"],
                "State": ["TX"],
                "Share (%)": [25.5],
                "Distribution": [100000.00],
            }
        )

        result = self.excel_service.validate_fund_source_headers(df)
        assert result is True
        assert len(self.excel_service.errors) == 0

    def test_validate_fund_source_headers_missing_headers(self):
        """Test fund source header validation with missing headers."""
        df = pd.DataFrame(
            {
                "CompanyName": ["TechCorp"],  # Wrong header
                "State": ["TX"],
                "Percentage": [25.5],  # Wrong header, missing Distribution
            }
        )

        result = self.excel_service.validate_fund_source_headers(df)
        assert result is False
        assert (
            len(self.excel_service.errors) == 3
        )  # Missing Company, Share (%), and Distribution

        error_messages = [error.error_message for error in self.excel_service.errors]
        assert any("Company" in msg for msg in error_messages)
        assert any("Share (%)" in msg for msg in error_messages)
        assert any("Distribution" in msg for msg in error_messages)

    def test_validate_fund_source_headers_normalized(self):
        """Test fund source header validation with whitespace normalization."""
        df = pd.DataFrame(
            {
                "  Company  ": ["TechCorp"],
                " State ": ["TX"],
                "Share   (%)": [25.5],
                "Distribution ": [100000.00],
            }
        )

        result = self.excel_service.validate_fund_source_headers(df)
        assert result is True
        assert len(self.excel_service.errors) == 0

    def test_validate_fund_source_row_valid(self):
        """Test fund source row validation with valid data."""
        row_data = {
            "Company": "TechCorp Inc",
            "State": "TX",
            "Share (%)": "45.5",
            "Distribution": "500000.00",
        }

        result = self.excel_service.validate_fund_source_row(row_data, 2)
        assert result is True
        assert len(self.excel_service.errors) == 0

    def test_validate_fund_source_row_empty_company(self):
        """Test fund source row validation with empty company name."""
        row_data = {
            "Company": "",
            "State": "TX",
            "Share (%)": "45.5",
            "Distribution": "500000.00",
        }

        result = self.excel_service.validate_fund_source_row(row_data, 2)
        assert result is False
        assert len(self.excel_service.errors) == 1
        assert self.excel_service.errors[0].error_code == "EMPTY_COMPANY_NAME"
        assert self.excel_service.errors[0].row_number == 2

    def test_validate_fund_source_row_invalid_state(self):
        """Test fund source row validation with invalid state code."""
        row_data = {
            "Company": "TechCorp Inc",
            "State": "ZZ",  # Invalid state
            "Share (%)": "45.5",
            "Distribution": "500000.00",
        }

        result = self.excel_service.validate_fund_source_row(row_data, 2)
        assert result is False
        assert len(self.excel_service.errors) == 1
        assert self.excel_service.errors[0].error_code == "INVALID_STATE_CODE"
        assert self.excel_service.errors[0].field_value == "ZZ"

    def test_validate_fund_source_row_invalid_share_percentage(self):
        """Test fund source row validation with invalid share percentage."""
        test_cases = [
            ("invalid", "INVALID_SHARE_PERCENTAGE"),
            ("120.0", "SHARE_PERCENTAGE_OUT_OF_RANGE"),  # > 100
            ("-5.0", "SHARE_PERCENTAGE_OUT_OF_RANGE"),  # < 0
            ("", "INVALID_SHARE_PERCENTAGE"),
        ]

        for share_value, expected_error in test_cases:
            self.excel_service.errors = []  # Reset errors
            row_data = {
                "Company": "TechCorp Inc",
                "State": "TX",
                "Share (%)": share_value,
                "Distribution": "500000.00",
            }

            result = self.excel_service.validate_fund_source_row(row_data, 2)
            assert result is False
            assert len(self.excel_service.errors) == 1
            assert self.excel_service.errors[0].error_code == expected_error

    def test_validate_fund_source_row_invalid_distribution_amount(self):
        """Test fund source row validation with invalid distribution amount."""
        test_cases = [
            ("0.00", "INVALID_DISTRIBUTION_AMOUNT"),
            ("-100000.00", "INVALID_DISTRIBUTION_AMOUNT"),
            ("", "INVALID_DISTRIBUTION_AMOUNT"),
        ]

        for dist_value, expected_error in test_cases:
            self.excel_service.errors = []  # Reset errors
            row_data = {
                "Company": "TechCorp Inc",
                "State": "TX",
                "Share (%)": "45.5",
                "Distribution": dist_value,
            }

            result = self.excel_service.validate_fund_source_row(row_data, 2)
            assert result is False
            assert len(self.excel_service.errors) >= 1
            error_codes = [error.error_code for error in self.excel_service.errors]
            assert expected_error in error_codes

    def test_parse_fund_source_row_valid(self):
        """Test parsing valid fund source row data."""
        row_data = {
            "Company": "  TechCorp Inc  ",
            "State": "  tx  ",  # Should be normalized to uppercase
            "Share (%)": "45.50%",  # Should handle % symbol
            "Distribution": "500,000.00",  # Should handle thousands separator
        }

        result = self.excel_service.parse_fund_source_row(row_data, 2)

        expected = {
            "company_name": "TechCorp Inc",
            "state_jurisdiction": "TX",
            "fund_share_percentage": Decimal("45.50"),
            "total_distribution_amount": Decimal("500000.00"),
            "row_number": 2,
        }

        assert result == expected

    def test_parse_fund_source_data_valid_file(self):
        """Test parsing fund source data from valid Excel file."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund A_Q1 2025 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        from openpyxl import load_workbook
        workbook = load_workbook(filename=test_file, data_only=True)
        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(workbook)
        )

        assert len(fund_source_errors) == 0
        assert len(fund_source_data) == 4  # Based on our test data

        # Check first record
        first_record = fund_source_data[0]
        assert first_record["company_name"] == "TechCorp Inc"
        assert first_record["state_jurisdiction"] == "TX"
        assert first_record["fund_share_percentage"] == Decimal("45.5")
        assert first_record["total_distribution_amount"] == Decimal("500000.00")

    def test_parse_fund_source_data_invalid_file(self):
        """Test parsing fund source data from invalid Excel file."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund B_Q2 2025 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        # Reset errors before test to get clean state
        self.excel_service.errors = []

        from openpyxl import load_workbook
        workbook = load_workbook(filename=test_file, data_only=True)
        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(workbook)
        )

        # Check combined errors from both service errors and fund_source_errors
        all_errors = self.excel_service.errors + fund_source_errors
        assert len(all_errors) > 0
        assert len(fund_source_data) < 4  # Some records should fail validation

        # Check for expected error types in combined errors
        all_error_codes = [error.error_code for error in all_errors]
        # Should have validation errors like SHARE_PERCENTAGE_OUT_OF_RANGE, INVALID_STATE_CODE, etc.
        expected_error_codes = [
            "SHARE_PERCENTAGE_OUT_OF_RANGE",
            "INVALID_STATE_CODE",
            "INVALID_DISTRIBUTION_AMOUNT",
            "DUPLICATE_COMPANY_STATE",
        ]
        assert any(error_code in all_error_codes for error_code in expected_error_codes)

    def test_parse_fund_source_data_missing_headers(self):
        """Test parsing fund source data with missing headers."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund C_Q3 2025 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        # Reset errors before test
        self.excel_service.errors = []

        from openpyxl import load_workbook
        workbook = load_workbook(filename=test_file, data_only=True)
        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(workbook)
        )

        # Check combined errors from both sources
        all_errors = self.excel_service.errors + fund_source_errors
        assert len(all_errors) > 0
        assert len(fund_source_data) == 0  # No valid data due to header issues

        # The error should be either in service errors or fund_source_errors
        all_error_codes = [error.error_code for error in all_errors]
        expected_codes = ["MISSING_FUND_SOURCE_HEADER", "FUND_SOURCE_PARSING_ERROR"]
        assert any(error_code in all_error_codes for error_code in expected_codes)

    def test_parse_fund_source_data_no_second_sheet(self):
        """Test parsing fund source data from file without second sheet."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund D_Q4 2025 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        from openpyxl import load_workbook
        workbook = load_workbook(filename=test_file, data_only=True)
        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(workbook)
        )

        # Should return empty data without errors (backward compatibility)
        assert len(fund_source_errors) == 0
        assert len(fund_source_data) == 0

    def test_parse_fund_source_data_empty_sheet(self):
        """Test parsing fund source data from empty second sheet."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund E_Q1 2024 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        from openpyxl import load_workbook
        workbook = load_workbook(filename=test_file, data_only=True)
        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(workbook)
        )

        # Should return empty data without errors
        assert len(fund_source_errors) == 0
        assert len(fund_source_data) == 0

    def test_parse_fund_source_data_duplicate_detection(self):
        """Test duplicate company/state combination detection."""
        # Create a mock workbook with duplicate data
        from unittest.mock import Mock
        from openpyxl import Workbook

        # Create a real workbook and modify it
        workbook = Workbook()
        if len(workbook.worksheets) > 1:
            del workbook.worksheets[1]  # Remove extra sheets
        workbook.create_sheet("Fund Source Data")  # Add our sheet
        sheet = workbook.worksheets[1]

        # Add headers
        sheet['A1'] = "Company"
        sheet['B1'] = "State"
        sheet['C1'] = "Share (%)"
        sheet['D1'] = "Distribution"

        # Add duplicate data
        sheet['A2'] = "TechCorp Inc"
        sheet['B2'] = "TX"
        sheet['C2'] = 45.5
        sheet['D2'] = 500000.00

        sheet['A3'] = "TechCorp Inc"  # Duplicate
        sheet['B3'] = "TX"            # Duplicate
        sheet['C3'] = 25.0
        sheet['D3'] = 300000.00

        sheet['A4'] = "MedDevice LLC"
        sheet['B4'] = "CA"
        sheet['C4'] = 30.0
        sheet['D4'] = 200000.00

        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(workbook)
        )

        assert len(fund_source_errors) == 1
        assert fund_source_errors[0].error_code == "DUPLICATE_COMPANY_STATE"
        assert "TechCorp Inc/TX" in fund_source_errors[0].error_message
        assert len(fund_source_data) == 2  # Only non-duplicate records

    def test_parse_fund_source_data_exception_handling(self):
        """Test exception handling in fund source data parsing."""
        from unittest.mock import Mock

        # Create a mock workbook that will raise an exception
        mock_workbook = Mock()
        mock_workbook.worksheets = []  # This will cause an IndexError

        fund_source_data, fund_source_errors = (
            self.excel_service.parse_fund_source_data(mock_workbook)
        )

        assert len(fund_source_data) == 0
        assert len(fund_source_errors) == 0  # IndexError is handled gracefully for missing sheets

    def test_fund_source_data_integration_with_main_parsing(self):
        """Test fund source data integration with main Excel parsing."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund A_Q1 2025 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        result = self.excel_service.parse_excel_file(test_file, test_file.name)

        assert result.fund_source_data is not None
        assert len(result.fund_source_data) > 0
        assert result.fund_source_errors is not None

        # Check that fund source data is included in the result
        first_fund_source = result.fund_source_data[0]
        assert "company_name" in first_fund_source
        assert "state_jurisdiction" in first_fund_source
        assert "fund_share_percentage" in first_fund_source
        assert "total_distribution_amount" in first_fund_source

    def test_fund_source_data_integration_errors_combined(self):
        """Test that fund source errors are combined with main parsing errors."""
        test_file = (
            self.test_data_dir
            / "(Input Data) Fund B_Q2 2025 distribution data_v1.3.xlsx"
        )

        if not test_file.exists():
            pytest.skip("Test file not found")

        result = self.excel_service.parse_excel_file(test_file, test_file.name)

        # Since the main sheet is valid, we might only have fund source errors
        # The important thing is that all errors are combined in result.errors
        all_error_codes = [error.error_code for error in result.errors]

        # Check that we have some validation errors from fund source parsing
        expected_error_codes = [
            "SHARE_PERCENTAGE_OUT_OF_RANGE",
            "INVALID_STATE_CODE",
            "INVALID_DISTRIBUTION_AMOUNT",
            "DUPLICATE_COMPANY_STATE",
        ]

        # Should have at least one of the expected error types
        assert any(error_code in all_error_codes for error_code in expected_error_codes)
        assert len(result.errors) > 0
